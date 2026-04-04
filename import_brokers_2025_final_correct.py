import psycopg2
import openpyxl
from datetime import datetime, timedelta
import os

DATABASE_URL = "postgresql://postgres:OMxLodDbSGnDJUQGVIkXSXYxfiRwQqFo@shortline.proxy.rlwy.net:45408/railway"

def normalize_price(value):
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip().replace('€', '').replace(' ', '')
    value_str = value_str.replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0

def parse_date(date_value):
    if date_value is None or date_value == '':
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    if hasattr(date_value, 'date'):
        return date_value.date()
    date_str = str(date_value).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(date_str, fmt).date()
        except:
            continue
    return None

# Lista de nomes de brokers conhecidos
BROKER_NAMES = [
    'ABBYCAR-POA', 'ABBYCAR-PREPAID', 'AP', 'API', 'API-WEB',
    'CARJET-PREPAID', 'CARJET-POA', 'DISCOVERCARS-POA', 'DISCOVERCARS-PREPAID',
    'RENTALCARS', 'VIP CARS-POA', 'VIP CARS', 'BROKERS - DIRECTOS'
]

try:
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    
    print("\n=== IMPORTANDO BROKERS 2025 ===\n")
    
    excel_dir = "./CM-25"
    files = sorted([f for f in os.listdir(excel_dir) if f.startswith('CM-') and f.endswith('.xlsx') and not f.startswith('~$')])
    
    total_imported = 0
    total_skipped = 0
    voucher_counter = 1
    
    for filename in files:
        filepath = os.path.join(excel_dir, filename)
        print(f"A processar {filename}...")
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        current_broker = None
        file_imported = 0
        file_skipped = 0
        
        for row_idx in range(1, ws.max_row + 1):
            col_a = ws.cell(row_idx, 1).value  # Voucher ou nome do broker
            col_b = ws.cell(row_idx, 2).value  # Data Entrega
            col_c = ws.cell(row_idx, 3).value  # Dias
            col_d = ws.cell(row_idx, 4).value  # Loyalty Card (preço)
            
            # Ignorar cabeçalho
            if col_a and 'VOUCHER' in str(col_a).upper():
                continue
            
            # Verificar se é linha de broker (nome na col A, sem data na col B)
            if col_a and not col_b:
                broker_name = str(col_a).strip()
                # Verificar se é um broker conhecido
                if broker_name in BROKER_NAMES:
                    current_broker = broker_name
                    print(f"  Broker: {current_broker}")
                continue
            
            # Se temos broker atual e linha tem data = é reserva
            if current_broker and col_b:
                voucher = None
                if col_a and str(col_a).strip():
                    voucher_str = str(col_a).strip()
                    # Se for número, é voucher
                    if voucher_str.replace('-', '').replace('.', '').isdigit():
                        voucher = voucher_str
                
                pickup_date = parse_date(col_b)
                days = int(col_c) if col_c and isinstance(col_c, (int, float)) else None
                price = normalize_price(col_d)
                
                if not pickup_date or price <= 0:
                    file_skipped += 1
                    continue
                
                # Inserir em broker_bookings (apenas campos necessários)
                try:
                    cur.execute("""
                        INSERT INTO broker_bookings (
                            voucher_number, broker_name, pickup_date, days, 
                            total_price, status, created_at
                        ) VALUES (
                            %s, %s, %s, %s, %s, 'confirmed', NOW()
                        )
                    """, (
                        voucher,
                        current_broker,
                        pickup_date,
                        days,
                        price
                    ))
                    file_imported += 1
                except Exception as e:
                    if 'unique constraint' not in str(e).lower():
                        print(f"  ✗ Erro linha {row_idx}: {e}")
                    file_skipped += 1
        
        conn.commit()
        total_imported += file_imported
        total_skipped += file_skipped
        print(f"  ✓ Importados: {file_imported}, Ignorados: {file_skipped}")
        wb.close()
    
    cur.close()
    conn.close()
    
    print(f"\n=== RESUMO ===")
    print(f"Total de brokers importados: {total_imported}")
    print(f"Total ignorado: {total_skipped}")
    print("\n✓ Importação concluída!")
    
except Exception as e:
    print(f"✗ Erro: {e}")
    import traceback
    traceback.print_exc()
