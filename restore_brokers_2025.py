import psycopg2
import openpyxl
from datetime import datetime
import os

DATABASE_URL = "postgresql://postgres:OMxLodDbSGnDJUQGVIkXSXYxfiRwQqFo@shortline.proxy.rlwy.net:45408/railway"

def normalize_price(value):
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip().replace('€', '').replace(' ', '')
    value_str = value_str.replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0

def parse_date(date_value):
    if date_value is None or date_value == '':
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    date_str = str(date_value).strip()
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(date_str, fmt).date()
        except:
            continue
    return None

# Lista de brokers conhecidos (não são comissionistas)
BROKER_NAMES = [
    'AP', 'CARJET-PREPAID', 'API-WEB', 'ABBYCAR-PREPAID', 'VIP CARS-POA',
    'ABBYCAR-POA', 'DISCOVERCARS-PREPAID', 'DISCOVERCARS-POA', 'RENTALCARS',
    'API', 'CARJET-POA', 'DISCOVERCARS', 'CARJET'
]

try:
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    
    print("\n=== RESTAURANDO DADOS DE BROKERS 2025 ===\n")
    
    excel_dir = "./CM-25"
    files = sorted([f for f in os.listdir(excel_dir) if f.startswith('CM-') and f.endswith('.xlsx') and not f.startswith('~$')])
    
    total_imported = 0
    total_skipped = 0
    voucher_counter = 1
    
    for filename in files:
        filepath = os.path.join(excel_dir, filename)
        print(f"A processar {filename}...")
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        current_broker = None
        file_imported = 0
        file_skipped = 0
        
        for row_idx in range(1, ws.max_row + 1):
            col1_value = ws.cell(row_idx, 1).value
            col2_value = ws.cell(row_idx, 2).value
            
            # Ignorar cabeçalho
            if col1_value and 'VOUCHER' in str(col1_value).upper():
                continue
            
            # Verificar se é linha de broker (nome na coluna 1, vazio na coluna 2)
            if col1_value and not col2_value:
                broker_name = str(col1_value).strip().upper()
                # Verificar se é um broker conhecido
                if any(b.upper() in broker_name or broker_name in b.upper() for b in BROKER_NAMES):
                    current_broker = broker_name
                else:
                    current_broker = None  # É comissionista, não broker
                continue
            
            # Se temos broker atual e linha tem data, importar para broker_bookings
            if col2_value and current_broker:
                voucher = None
                if col1_value and str(col1_value).strip():
                    voucher_str = str(col1_value).strip()
                    if voucher_str.replace('-', '').isdigit():
                        voucher = voucher_str
                
                pickup_date = parse_date(col2_value)
                days = ws.cell(row_idx, 3).value
                price = normalize_price(ws.cell(row_idx, 4).value)
                
                if not pickup_date or price <= 0:
                    file_skipped += 1
                    continue
                
                # Gerar voucher se não existir
                if not voucher:
                    voucher = f"BRK-2025-{voucher_counter:06d}"
                    voucher_counter += 1
                
                # Calcular dropoff_date
                from datetime import timedelta
                dropoff_date = pickup_date
                if days and isinstance(days, (int, float)) and days > 0:
                    dropoff_date = pickup_date + timedelta(days=int(days))
                
                # Inserir em broker_bookings
                try:
                    cur.execute("""
                        INSERT INTO broker_bookings (
                            voucher_number, broker_name, client_name, client_email, 
                            client_phone, pickup_date, pickup_time, dropoff_date, 
                            dropoff_time, price, status, created_at, updated_at,
                            pickup_location, dropoff_location, vehicle_group
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'confirmed',
                            NOW(), NOW(), 'N/A', 'N/A', 'N/A'
                        )
                    """, (
                        voucher,
                        current_broker,
                        'N/A',
                        'broker@rentalprices.pt',
                        'N/A',
                        pickup_date,
                        '10:00',
                        dropoff_date,
                        '10:00',
                        price
                    ))
                    file_imported += 1
                except Exception as e:
                    if 'unique constraint' not in str(e).lower():
                        print(f"  ✗ Erro: {e}")
                    file_skipped += 1
        
        conn.commit()
        total_imported += file_imported
        total_skipped += file_skipped
        print(f"  ✓ Importados: {file_imported}, Ignorados: {file_skipped}")
        wb.close()
    
    cur.close()
    conn.close()
    
    print(f"\n=== RESUMO ===")
    print(f"Total de brokers importados: {total_imported}")
    print(f"Total ignorado: {total_skipped}")
    print("\n✓ Restauração concluída!")
    
except Exception as e:
    print(f"✗ Erro: {e}")
    import traceback
    traceback.print_exc()
