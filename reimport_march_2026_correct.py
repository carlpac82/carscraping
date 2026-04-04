import openpyxl
import psycopg2
from datetime import datetime, timedelta

DATABASE_URL = "postgresql://postgres:OMxLodDbSGnDJUQGVIkXSXYxfiRwQqFo@shortline.proxy.rlwy.net:45408/railway"

def parse_date(date_value):
    """Parse date from Excel"""
    if date_value is None or date_value == '':
        return None
    if isinstance(date_value, datetime):
        return date_value
    if hasattr(date_value, 'date'):
        return date_value
    date_str = str(date_value).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(date_str, fmt)
        except:
            continue
    return None

def get_commissioners_map(conn):
    """Obter mapa de comissionistas: nome -> (id, commission_rate)"""
    cur = conn.cursor()
    cur.execute("SELECT id, name, commission_rate FROM commissioners")
    commissioners = {}
    for row in cur.fetchall():
        name_normalized = row[1].strip().upper()
        commissioners[name_normalized] = (row[0], float(row[2]) if row[2] else 0.15)
    cur.close()
    return commissioners

def find_commissioner(broker_name, commissioners_map):
    """Encontrar ID e taxa do comissionista pelo nome"""
    if not broker_name:
        return None, None
    
    broker_normalized = broker_name.strip().upper()
    
    if broker_normalized in commissioners_map:
        return commissioners_map[broker_normalized]
    
    for comm_name, (comm_id, comm_rate) in commissioners_map.items():
        if comm_name in broker_normalized or broker_normalized in comm_name:
            return comm_id, comm_rate
    
    return None, None

def reimport_march_2026():
    """Reimportar dados de março 2026 corretamente"""
    
    file_path = '/Users/filipepacheco/CascadeProjects/carscraping/CM-03-2026.xlsx'
    
    print("=" * 80)
    print("REIMPORTAR MARÇO 2026 - CORRETO")
    print("=" * 80)
    
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    
    # Obter mapa de comissionistas
    commissioners_map = get_commissioners_map(conn)
    print(f"\n📋 Comissionistas na BD: {len(commissioners_map)}")
    
    # Abrir ficheiro Excel
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"📄 Ficheiro: {file_path}")
    print(f"📊 Total de linhas: {ws.max_row}")
    
    current_broker = None
    commissioner_id = None
    commission_rate = None
    
    imported_count = 0
    ignored_count = 0
    
    print("\n" + "=" * 80)
    print("PROCESSANDO LINHAS:")
    print("=" * 80)
    
    for row_num in range(1, ws.max_row + 1):
        col_a = ws.cell(row=row_num, column=1).value  # Voucher ou Broker
        col_b = ws.cell(row=row_num, column=2).value  # Data Entrega
        col_c = ws.cell(row=row_num, column=3).value  # Dias
        col_d = ws.cell(row=row_num, column=4).value  # Price (Loyalty Card)
        
        # Linha de cabeçalho - ignorar
        if row_num == 1:
            continue
        
        # Linha de broker: tem texto em A mas não tem data em B
        if col_a and not col_b:
            current_broker = str(col_a).strip()
            commissioner_id, commission_rate = find_commissioner(current_broker, commissioners_map)
            
            if commissioner_id:
                print(f"\n🏨 {current_broker} (ID: {commissioner_id}, Taxa: {commission_rate*100:.0f}%)")
            else:
                print(f"\n⚠️  {current_broker} - NÃO ENCONTRADO NA BD")
            continue
        
        # Linha de booking: tem data em B
        if col_b and current_broker and commissioner_id:
            # Parse da data
            pickup_datetime = parse_date(col_b)
            if not pickup_datetime:
                print(f"  ⚠️  Linha {row_num}: Data inválida")
                ignored_count += 1
                continue
            
            # Parse do preço (coluna D - Loyalty Card)
            price = 0
            if col_d:
                try:
                    price = float(col_d)
                except:
                    print(f"  ⚠️  Linha {row_num}: Preço inválido")
                    ignored_count += 1
                    continue
            
            if price <= 0:
                print(f"  ⚠️  Linha {row_num}: Preço zero ou negativo")
                ignored_count += 1
                continue
            
            # Parse dos dias
            days = 1
            if col_c:
                try:
                    days = int(col_c)
                except:
                    days = 1
            
            # Voucher (se existir na coluna A)
            voucher = None
            if col_a:
                voucher_str = str(col_a).strip()
                if voucher_str and voucher_str.upper() != current_broker.upper():
                    voucher = voucher_str
            
            # Calcular dropoff_date
            pickup_date = pickup_datetime.date()
            pickup_time = pickup_datetime.strftime('%H:%M')
            dropoff_date = pickup_date + timedelta(days=days)
            
            # Calcular comissão: price × commission_rate
            commission_amount = price * commission_rate
            
            # Inserir na BD
            try:
                insert_query = """
                    INSERT INTO commission_bookings (
                        commissioner_id, voucher_number, client_name, client_email, client_phone,
                        pickup_date, pickup_time, dropoff_date, dropoff_time,
                        pickup_location, dropoff_location, vehicle_group, extras,
                        price, commission_amount, status, created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                        CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
                    )
                """
                
                cur.execute(insert_query, (
                    commissioner_id, voucher, 'Commission Client', 'commission@rentalprices.pt', '',
                    pickup_date, pickup_time, dropoff_date, '00:00',
                    '', '', '', '[]',
                    price, commission_amount, 'confirmed'
                ))
                
                voucher_str = f" [{voucher}]" if voucher else ""
                print(f"  ✅ {pickup_date}{voucher_str} - {days}d - €{price:.2f} (comissão: €{commission_amount:.2f})")
                imported_count += 1
                
            except Exception as e:
                print(f"  ❌ Erro ao inserir linha {row_num}: {str(e)}")
                ignored_count += 1
    
    conn.commit()
    
    print("\n" + "=" * 80)
    print("RESUMO DA IMPORTAÇÃO:")
    print("=" * 80)
    print(f"✅ Reservas importadas: {imported_count}")
    print(f"⚠️  Reservas ignoradas: {ignored_count}")
    
    # Verificar totais
    query = """
        SELECT 
            COUNT(*) as count,
            SUM(price) as total_price,
            SUM(commission_amount) as total_commission
        FROM commission_bookings
        WHERE EXTRACT(YEAR FROM pickup_date) = 2026
        AND EXTRACT(MONTH FROM pickup_date) = 3
    """
    
    cur.execute(query)
    row = cur.fetchone()
    
    print(f"\n📊 VERIFICAÇÃO NA BD:")
    print(f"  - Total de reservas: {row[0]}")
    print(f"  - Total price: €{float(row[1]) if row[1] else 0:.2f}")
    print(f"  - Total comissão: €{float(row[2]) if row[2] else 0:.2f}")
    
    cur.close()
    conn.close()

if __name__ == "__main__":
    reimport_march_2026()
