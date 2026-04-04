import openpyxl
import os

excel_dir = "./CM-25"
files = sorted([f for f in os.listdir(excel_dir) if f.startswith('CM-') and f.endswith('.xlsx') and not f.startswith('~$')])

for filename in files:
    filepath = os.path.join(excel_dir, filename)
    print(f"\n{'='*80}")
    print(f"{filename}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    brokers_found = []
    
    # Analisar TODAS as linhas
    for row_idx in range(1, ws.max_row + 1):
        col1 = ws.cell(row_idx, 1).value
        col2 = ws.cell(row_idx, 2).value
        
        # Se tem valor na col1 mas não na col2, pode ser broker ou comissionista
        if col1 and not col2:
            col1_str = str(col1).strip()
            # Ignorar cabeçalho
            if 'VOUCHER' not in col1_str.upper():
                # Verificar se é broker (maiúsculas, hífen, contém AP/API/CARJET/etc)
                broker_keywords = ['AP', 'API', 'CARJET', 'ABBYCAR', 'DISCOVERCARS', 'VIP', 'RENTALCARS', 'BROKERS']
                is_broker = any(keyword in col1_str.upper() for keyword in broker_keywords)
                
                if is_broker:
                    if col1_str not in brokers_found:
                        brokers_found.append(col1_str)
                        print(f"Linha {row_idx}: '{col1_str}'")
    
    print(f"\nTotal de brokers únicos: {len(brokers_found)}")
    wb.close()
