"""
Endpoint para exportar dados de comissionistas e brokers para Excel
Adicionar ao main.py após o endpoint /api/admin/commissions/print-pdf
"""

@app.get("/api/admin/commissions/export-excel")
async def admin_commissions_export_excel(request: Request):
    """Generate Excel report of commissions and brokers for selected month/year"""
    try:
        require_admin(request)
    except HTTPException:
        return JSONResponse({"ok": False, "error": "Unauthorized"}, status_code=403)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from datetime import datetime
        
        # Get query parameters
        month = request.query_params.get("month", "")
        year = request.query_params.get("year", "")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # ===================================================================
        # SEPARADOR 1: COMISSIONISTAS
        # ===================================================================
        ws_comm = wb.create_sheet("COMISSIONISTAS")
        
        # Header style
        header_fill = PatternFill(start_color="009cb6", end_color="009cb6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Set headers
        ws_comm['A1'] = 'Voucher'
        ws_comm['B1'] = 'Data Entrega'
        ws_comm['C1'] = 'Dias'
        ws_comm['D1'] = 'Base'
        ws_comm['E1'] = 'Comissão'
        
        # Apply header style
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws_comm[f'{col}1']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Fetch commissioners data
        with _db_lock:
            con = _db_connect()
            try:
                if USE_POSTGRES:
                    query = """
                        SELECT 
                            c.name as commissioner_name,
                            cb.voucher_number,
                            cb.pickup_date,
                            EXTRACT(DAY FROM (cb.dropoff_date - cb.pickup_date)) as days,
                            cb.base_price,
                            cb.commission_amount
                        FROM commission_bookings cb
                        JOIN commissioners c ON cb.commissioner_id = c.id
                        WHERE 1=1
                    """
                    params = []
                    
                    if month:
                        query += " AND EXTRACT(MONTH FROM cb.pickup_date) = %s"
                        params.append(int(month))
                    
                    if year:
                        query += " AND EXTRACT(YEAR FROM cb.pickup_date) = %s"
                        params.append(int(year))
                    
                    query += " ORDER BY c.name, cb.pickup_date"
                    
                    cur = con.cursor()
                    cur.execute(query, params)
                    commissioners_data = cur.fetchall()
                else:
                    query = """
                        SELECT 
                            c.name as commissioner_name,
                            cb.voucher_number,
                            cb.pickup_date,
                            CAST((julianday(cb.dropoff_date) - julianday(cb.pickup_date)) AS INTEGER) as days,
                            cb.base_price,
                            cb.commission_amount
                        FROM commission_bookings cb
                        JOIN commissioners c ON cb.commissioner_id = c.id
                        WHERE 1=1
                    """
                    params = []
                    
                    if month:
                        query += " AND CAST(strftime('%m', cb.pickup_date) AS INTEGER) = ?"
                        params.append(int(month))
                    
                    if year:
                        query += " AND CAST(strftime('%Y', cb.pickup_date) AS INTEGER) = ?"
                        params.append(int(year))
                    
                    query += " ORDER BY c.name, cb.pickup_date"
                    
                    cur = con.execute(query, params)
                    commissioners_data = cur.fetchall()
                
            finally:
                con.close()
        
        # Group by commissioner
        current_row = 2
        current_commissioner = None
        commissioner_totals = {}
        
        for row_data in commissioners_data:
            commissioner_name = row_data[0]
            voucher = row_data[1] or ''
            pickup_date = row_data[2]
            days = row_data[3] or 1
            base_price = float(row_data[4]) if row_data[4] else 0
            commission = float(row_data[5]) if row_data[5] else 0
            
            # New commissioner - add header row
            if commissioner_name != current_commissioner:
                if current_commissioner:
                    current_row += 1  # Empty row between commissioners
                
                ws_comm[f'A{current_row}'] = commissioner_name
                ws_comm[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                current_commissioner = commissioner_name
                commissioner_totals[commissioner_name] = {'base': 0, 'commission': 0}
            
            # Add data row
            ws_comm[f'A{current_row}'] = voucher
            ws_comm[f'B{current_row}'] = pickup_date
            ws_comm[f'C{current_row}'] = days
            ws_comm[f'D{current_row}'] = base_price
            ws_comm[f'E{current_row}'] = commission
            
            # Format date
            ws_comm[f'B{current_row}'].number_format = 'DD/MM/YYYY HH:MM'
            
            # Format currency
            ws_comm[f'D{current_row}'].number_format = '€#,##0.00'
            ws_comm[f'E{current_row}'].number_format = '€#,##0.00'
            
            # Update totals
            commissioner_totals[commissioner_name]['base'] += base_price
            commissioner_totals[commissioner_name]['commission'] += commission
            
            current_row += 1
        
        # Adjust column widths
        ws_comm.column_dimensions['A'].width = 25
        ws_comm.column_dimensions['B'].width = 20
        ws_comm.column_dimensions['C'].width = 10
        ws_comm.column_dimensions['D'].width = 15
        ws_comm.column_dimensions['E'].width = 15
        
        # ===================================================================
        # SEPARADOR 2: AP+API-WEB+BROKERS
        # ===================================================================
        ws_brokers = wb.create_sheet("AP+API-WEB+BROKERS")
        
        # Set headers
        ws_brokers['A1'] = 'Voucher'
        ws_brokers['B1'] = 'Data Entrega'
        ws_brokers['C1'] = 'Dias'
        ws_brokers['D1'] = 'Base'
        
        # Apply header style
        for col in ['A', 'B', 'C', 'D']:
            cell = ws_brokers[f'{col}1']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Fetch brokers data
        with _db_lock:
            con = _db_connect()
            try:
                if USE_POSTGRES:
                    query = """
                        SELECT 
                            broker_name,
                            voucher_number,
                            pickup_date,
                            days,
                            total_price
                        FROM broker_bookings
                        WHERE 1=1
                    """
                    params = []
                    
                    if month:
                        query += " AND EXTRACT(MONTH FROM pickup_date) = %s"
                        params.append(int(month))
                    
                    if year:
                        query += " AND EXTRACT(YEAR FROM pickup_date) = %s"
                        params.append(int(year))
                    
                    query += " ORDER BY broker_name, pickup_date"
                    
                    cur = con.cursor()
                    cur.execute(query, params)
                    brokers_data = cur.fetchall()
                else:
                    query = """
                        SELECT 
                            broker_name,
                            voucher_number,
                            pickup_date,
                            days,
                            total_price
                        FROM broker_bookings
                        WHERE 1=1
                    """
                    params = []
                    
                    if month:
                        query += " AND CAST(strftime('%m', pickup_date) AS INTEGER) = ?"
                        params.append(int(month))
                    
                    if year:
                        query += " AND CAST(strftime('%Y', pickup_date) AS INTEGER) = ?"
                        params.append(int(year))
                    
                    query += " ORDER BY broker_name, pickup_date"
                    
                    cur = con.execute(query, params)
                    brokers_data = cur.fetchall()
                
            finally:
                con.close()
        
        # Group by broker
        current_row = 2
        current_broker = None
        broker_totals = {}
        
        for row_data in brokers_data:
            broker_name = row_data[0]
            voucher = row_data[1] or ''
            pickup_date = row_data[2]
            days = row_data[3] or 1
            total_price = float(row_data[4]) if row_data[4] else 0
            base_price = total_price / 1.23  # Remove IVA
            
            # New broker - add header row
            if broker_name != current_broker:
                if current_broker:
                    current_row += 1  # Empty row between brokers
                
                ws_brokers[f'A{current_row}'] = broker_name
                ws_brokers[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                current_broker = broker_name
                broker_totals[broker_name] = 0
            
            # Add data row
            ws_brokers[f'A{current_row}'] = voucher
            ws_brokers[f'B{current_row}'] = pickup_date
            ws_brokers[f'C{current_row}'] = days
            ws_brokers[f'D{current_row}'] = base_price
            
            # Format date
            ws_brokers[f'B{current_row}'].number_format = 'DD/MM/YYYY HH:MM'
            
            # Format currency
            ws_brokers[f'D{current_row}'].number_format = '€#,##0.00'
            
            # Update totals
            broker_totals[broker_name] += base_price
            
            current_row += 1
        
        # Adjust column widths
        ws_brokers.column_dimensions['A'].width = 25
        ws_brokers.column_dimensions['B'].width = 20
        ws_brokers.column_dimensions['C'].width = 10
        ws_brokers.column_dimensions['D'].width = 15
        
        # ===================================================================
        # SAVE AND RETURN
        # ===================================================================
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename
        month_name = datetime(2000, int(month) if month else 1, 1).strftime('%B') if month else 'Todos'
        year_str = year if year else 'Todos'
        filename = f"Comissoes_Brokers_{month_name}_{year_str}.xlsx"
        
        return Response(
            content=excel_file.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
        
    except Exception as e:
        import traceback
        print(f"Error generating Excel: {e}")
        traceback.print_exc()
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)
