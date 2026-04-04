import openpyxl
import os

# Analisar estrutura dos ficheiros de brokers 2026
excel_dir = "./CM-26"
files = sorted([f for f in os.listdir(excel_dir) if f.endswith('.xlsx') and not f.startswith('~$')])

for filename in files:
    filepath = os.path.join(excel_dir, filename)
    print(f"\n{'='*60}")
    print(f"Ficheiro: {filename}")
    print('='*60)
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    print(f"Total de linhas: {ws.max_row}")
    print(f"Total de colunas: {ws.max_column}")
    
    print("\nPrimeiras 20 linhas:")
    print("-" * 60)
    
    for row_idx in range(1, min(21, ws.max_row + 1)):
        col_a = ws.cell(row_idx, 1).value
        col_b = ws.cell(row_idx, 2).value
        col_c = ws.cell(row_idx, 3).value
        col_d = ws.cell(row_idx, 4).value
        
        print(f"Linha {row_idx:3d}: A={str(col_a)[:30]:30s} | B={str(col_b)[:15]:15s} | C={str(col_c)[:10]:10s} | D={str(col_d)[:15]:15s}")
    
    wb.close()
