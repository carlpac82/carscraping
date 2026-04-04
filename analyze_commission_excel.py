import openpyxl

filepath = "./2025/CM-04-2025.xlsx"

wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb.active

print(f"\n=== ANÁLISE DE {filepath} ===\n")
print(f"Total de linhas: {ws.max_row}")
print(f"Total de colunas: {ws.max_column}\n")

print("=== PRIMEIRAS 30 LINHAS ===\n")
for row_idx in range(1, min(31, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(7, ws.max_column + 1)):
        cell_value = ws.cell(row_idx, col_idx).value
        row_data.append(str(cell_value)[:30] if cell_value else "")
    print(f"Linha {row_idx:2d}: {' | '.join(row_data)}")

wb.close()
