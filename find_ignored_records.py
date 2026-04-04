import openpyxl

def normalize_price(value):
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip().replace('€', '').replace(' ', '')
    value_str = value_str.replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0

# Janeiro - 1 ignorado
print("\n=== JANEIRO (CM-01-2025.xlsx) ===\n")
wb = openpyxl.load_workbook("./2025/CM-01-2025.xlsx", data_only=True)
ws = wb.active

for row_idx in range(1, ws.max_row + 1):
    col1 = ws.cell(row_idx, 1).value
    col2 = ws.cell(row_idx, 2).value
    col4 = ws.cell(row_idx, 4).value
    
    if col2:  # Tem data
        price = normalize_price(col4)
        if price <= 0:
            print(f"Linha {row_idx}: Preço inválido = {col4}")
            print(f"  Col1: {col1}")
            print(f"  Col2: {col2}")
            print(f"  Col4: {col4}")

wb.close()

# Março - 1 ignorado
print("\n=== MARÇO (CM-03-2025.xlsx) ===\n")
wb = openpyxl.load_workbook("./2025/CM-03-2025.xlsx", data_only=True)
ws = wb.active

for row_idx in range(1, ws.max_row + 1):
    col1 = ws.cell(row_idx, 1).value
    col2 = ws.cell(row_idx, 2).value
    col4 = ws.cell(row_idx, 4).value
    
    if col2:  # Tem data
        price = normalize_price(col4)
        if price <= 0:
            print(f"Linha {row_idx}: Preço inválido = {col4}")
            print(f"  Col1: {col1}")
            print(f"  Col2: {col2}")
            print(f"  Col4: {col4}")

wb.close()
