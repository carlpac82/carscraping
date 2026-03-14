import openpyxl

wb = openpyxl.load_workbook('CARALLIANCE-ALBUFEIRA-ALL-PERIODS.xlsx')
ws = wb['Prices']

print('=== ANÁLISE DO FICHEIRO CARALLIANCE ===\n')

# Count rows
total_rows = ws.max_row - 1  # -1 for header
print(f'Total de linhas de dados: {total_rows}')
print(f'Total de colunas: {ws.max_column}\n')

# Show headers
print('Headers:')
for col_idx in range(1, ws.max_column + 1):
    header = ws.cell(1, col_idx).value
    print(f'  Col {col_idx}: {header}')

print('\n=== PRIMEIRAS 5 LINHAS DE DADOS ===\n')

for row_idx in range(2, min(7, ws.max_row + 1)):
    print(f'Linha {row_idx}:')
    service = ws.cell(row_idx, 1).value
    office = ws.cell(row_idx, 2).value
    date_from = ws.cell(row_idx, 3).value
    date_to = ws.cell(row_idx, 4).value
    print(f'  Service: {service}')
    print(f'  Office: {office}')
    print(f'  Período: {date_from} até {date_to}')
    
    # Show price values
    print(f'  Preços:', end='')
    for col_idx in range(5, min(10, ws.max_column + 1)):
        val = ws.cell(row_idx, col_idx).value
        if val is not None:
            try:
                print(f' {float(val):.2f}', end='')
            except:
                print(f' {val}', end='')
    print('\n')

print('\n=== DETALHES DA LINHA 29 ===\n')

row_idx = 29
print(f'ServiceName: {ws.cell(row_idx, 1).value}')
print(f'Office: {ws.cell(row_idx, 2).value}')
print(f'DateFrom: {ws.cell(row_idx, 3).value}')
print(f'DateTo: {ws.cell(row_idx, 4).value}')
print(f'\nTodos os preços desta linha:')
for col_idx in range(5, 17):
    header = ws.cell(1, col_idx).value
    value = ws.cell(row_idx, col_idx).value
    if value is not None:
        try:
            print(f'  {header}: {float(value):.2f}')
        except:
            print(f'  {header}: {value}')
    else:
        print(f'  {header}: (vazio)')
print(f'  DiscountPercentage: {ws.cell(row_idx, 17).value}')
