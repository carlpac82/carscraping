import psycopg2
import openpyxl
from datetime import datetime
import os

DATABASE_URL = "postgresql://postgres:OMxLodDbSGnDJUQGVIkXSXYxfiRwQqFo@shortline.proxy.rlwy.net:45408/railway"

def normalize_price(value):
    """Normalizar valores de preço"""
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip().replace('€', '').replace(' ', '')
    value_str = value_str.replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0

def parse_date(date_value):
    """Parse date from various formats"""
    if date_value is None or date_value == '':
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    date_str = str(date_value).strip()
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(date_str, fmt).date()
        except:
            continue
    return None

def get_commissioners_map(conn):
    """Obter mapa de comissionistas: nome -> (id, commission_rate)"""
    cur = conn.cursor()
    cur.execute("SELECT id, name, commission_rate FROM commissioners")
    commissioners = {}
    for row in cur.fetchall():
        name_normalized = row[1].strip().upper()
        commissioners[name_normalized] = (row[0], float(row[2]) if row[2] else 15.0)
    cur.close()
    return commissioners

def find_commissioner(broker_name, commissioners_map):
    """Encontrar ID e taxa do comissionista pelo nome"""
    if not broker_name:
        return None, None
    
    name_normalized = broker_name.strip().upper()
    
    # Procura exata
    if name_normalized in commissioners_map:
        return commissioners_map[name_normalized]
    
    # Procura parcial
    for comm_name, (comm_id, comm_rate) in commissioners_map.items():
        if name_normalized in comm_name or comm_name in name_normalized:
            return comm_id, comm_rate
    
    return None, None

try:
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    
    # Obter mapa de comissionistas
    print("A carregar comissionistas da base de dados...")
    commissioners_map = get_commissioners_map(conn)
    print(f"✓ {len(commissioners_map)} comissionistas carregados\n")
    
    # Limpar dados de 2025 em commission_bookings
    print("A limpar dados de 2025 em commission_bookings...")
    cur.execute("""
        DELETE FROM commission_bookings 
        WHERE EXTRACT(YEAR FROM pickup_date) = 2025
    """)
    deleted = cur.rowcount
    conn.commit()
    print(f"✓ {deleted} registos removidos\n")
    
    # Processar ficheiros Excel
    excel_dir = "./2025"
    files = sorted([f for f in os.listdir(excel_dir) if f.startswith('CM-') and f.endswith('.xlsx')])
    
    total_imported = 0
    total_skipped = 0
    not_found_commissioners = set()
    voucher_counter = 1
    
    for filename in files:
        filepath = os.path.join(excel_dir, filename)
        print(f"A processar {filename}...")
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        current_commissioner = None
        current_commissioner_id = None
        current_commission_rate = None
        file_imported = 0
        file_skipped = 0
        
        # Processar todas as linhas
        for row_idx in range(1, ws.max_row + 1):
            col1_value = ws.cell(row_idx, 1).value
            col2_value = ws.cell(row_idx, 2).value
            
            # Verificar se é linha de cabeçalho (ignorar)
            if col1_value and 'VOUCHER' in str(col1_value).upper():
                continue
            
            # Verificar se é linha de comissionista (nome na coluna 1, vazio na coluna 2)
            if col1_value and not col2_value:
                current_commissioner = str(col1_value).strip()
                current_commissioner_id, current_commission_rate = find_commissioner(current_commissioner, commissioners_map)
                
                if not current_commissioner_id:
                    not_found_commissioners.add(current_commissioner)
                continue
            
            # Verificar se é linha de dados (tem data na coluna 2)
            if col2_value and current_commissioner_id:
                pickup_date = parse_date(col2_value)
                days = ws.cell(row_idx, 3).value
                price = normalize_price(ws.cell(row_idx, 4).value)
                
                if not pickup_date or price <= 0:
                    file_skipped += 1
                    continue
                
                # Calcular dropoff_date se temos dias, senão usar pickup_date
                dropoff_date = pickup_date
                if days and isinstance(days, (int, float)) and days > 0:
                    from datetime import timedelta
                    dropoff_date = pickup_date + timedelta(days=int(days))
                
                # Calcular comissão
                commission_amount = price * (current_commission_rate / 100) if current_commission_rate else 0
                
                # Gerar voucher único
                voucher_unique = f"COMM-2025-{voucher_counter:06d}"
                voucher_counter += 1
                
                # Inserir em commission_bookings
                try:
                    cur.execute("""
                        INSERT INTO commission_bookings (
                            voucher_number, client_name, client_email, client_phone,
                            pickup_date, pickup_time, dropoff_date, dropoff_time,
                            price, commissioner_id, commission_rate, commission_amount, 
                            commission_paid, status, created_at, updated_at, 
                            pickup_location, dropoff_location, vehicle_group
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, false, 
                            'confirmed', NOW(), NOW(), 'N/A', 'N/A', 'N/A'
                        )
                    """, (
                        voucher_unique,
                        'Loyalty Card',
                        'loyalty@rentalprices.pt',
                        'N/A',
                        pickup_date,
                        '10:00',
                        dropoff_date,
                        '10:00',
                        price,
                        current_commissioner_id,
                        current_commission_rate,
                        commission_amount
                    ))
                    file_imported += 1
                except Exception as e:
                    print(f"  ✗ Erro ao inserir: {e}")
                    file_skipped += 1
        
        conn.commit()
        total_imported += file_imported
        total_skipped += file_skipped
        print(f"  ✓ Importados: {file_imported}, Ignorados: {file_skipped}")
        wb.close()
    
    cur.close()
    conn.close()
    
    print(f"\n=== RESUMO ===")
    print(f"Total importado: {total_imported}")
    print(f"Total ignorado: {total_skipped}")
    
    if not_found_commissioners:
        print(f"\n⚠ Comissionistas não encontrados na BD ({len(not_found_commissioners)}):")
        for name in sorted(not_found_commissioners):
            print(f"  - {name}")
    
    print("\n✓ Importação concluída!")
    
except Exception as e:
    print(f"✗ Erro: {e}")
    import traceback
    traceback.print_exc()
