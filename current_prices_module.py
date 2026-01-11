"""
Módulo de Gestão de Preços Atuais
Permite visualizar, editar, importar e exportar preços nos formatos:
- Abbycar
- Brokers (Albufeira e Faro)
- Website
"""
import json
import os
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

COMMISSION_RATE = 0.1366  # 13.66%
# TODOS os grupos (incluindo K, comerciais e X)
GRUPOS = ['B1', 'B2', 'BK1', 'BK2', 'C', 'C3', 'C4', 'C5', 'D', 'DK', 'E1', 'E2', 'EK1', 'EK2', 'F', 'FK', 'G', 'J1', 'J2', 'JK1', 'JK2', 'L1', 'L2', 'LK1', 'M1', 'M2', 'MK1', 'MK2', 'N', 'NK', 'X']
GRUPOS_K = ['BK1', 'BK2', 'DK', 'EK1', 'EK2', 'FK', 'JK1', 'JK2', 'LK1', 'MK1', 'MK2', 'NK']
GRUPOS_COMERCIAIS = ['C3', 'C4', 'C5']
DIAS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 14, 22, 28, 31, 60]

def calculate_commission(net_price):
    """Calcula preço com comissão"""
    return round(net_price * (1 + COMMISSION_RATE), 2)

def load_prices_from_db(conn, location, month, year, day_start=None, day_end=None):
    """Carrega preços da base de dados
    Se day_start/day_end não forem especificados, retorna todos os períodos do mês
    """
    try:
        # Detectar se é PostgreSQL ou SQLite
        is_postgres = conn.__class__.__module__ in ['psycopg2.extensions', 'psycopg2._psycopg']
        
        if is_postgres:
            # PostgreSQL
            with conn.cursor() as cur:
                if day_start is not None and day_end is not None:
                    # Carregar período específico
                    cur.execute("""
                        SELECT prices_data, updated_at, day_start, day_end
                        FROM current_prices 
                        WHERE location = %s AND month = %s AND year = %s 
                          AND day_start = %s AND day_end = %s
                    """, (location, month, year, day_start, day_end))
                    
                    row = cur.fetchone()
                    if row and row[0]:
                        prices = json.loads(row[0])
                        updated_at = row[1].isoformat() if row[1] else None
                        logging.info(f"✅ Preços carregados: {location}, mês {month}/{year}, dias {day_start}-{day_end}")
                        return prices, updated_at
                    else:
                        logging.info(f"ℹ️ Sem preços para: {location}, mês {month}/{year}, dias {day_start}-{day_end}")
                        return None, None
                else:
                    # Carregar todos os períodos do mês
                    logging.info(f"Loading all periods for {location}, month={month}, year={year}")
                    cur.execute("""
                        SELECT prices_data, updated_at, day_start, day_end
                        FROM current_prices 
                        WHERE location = %s AND month = %s AND year = %s
                        ORDER BY day_start
                    """, (location, month, year))
                    
                    rows = cur.fetchall()
                    logging.info(f"Found {len(rows)} rows for {location}, {month}/{year}")
                    
                    periods = []
                    for row in rows:
                        # Converter datetime para string se necessário
                        updated_at_str = None
                        if row[1]:
                            if hasattr(row[1], 'isoformat'):
                                updated_at_str = row[1].isoformat()
                            else:
                                updated_at_str = str(row[1])
                        
                        periods.append({
                            'prices': json.loads(row[0]),
                            'updated_at': updated_at_str,
                            'day_start': row[2] if row[2] is not None else 1,
                            'day_end': row[3] if row[3] is not None else 31
                        })
                    
                    if periods:
                        logging.info(f"✅ {len(periods)} período(s) carregado(s): {location}, mês {month}/{year}")
                    else:
                        logging.info(f"ℹ️ Sem preços para: {location}, mês {month}/{year}")
                    
                    return periods, None
        else:
            # SQLite
            if day_start is not None and day_end is not None:
                # Carregar período específico
                cursor = conn.execute("""
                    SELECT prices_data, updated_at, day_start, day_end
                    FROM current_prices 
                    WHERE location = ? AND month = ? AND year = ? 
                      AND day_start = ? AND day_end = ?
                """, (location, month, year, day_start, day_end))
                
                row = cursor.fetchone()
                if row and row[0]:
                    prices = json.loads(row[0])
                    updated_at = row[1] if row[1] else None
                    logging.info(f"✅ Preços carregados: {location}, mês {month}/{year}, dias {day_start}-{day_end}")
                    return prices, updated_at
                else:
                    logging.info(f"ℹ️ Sem preços para: {location}, mês {month}/{year}, dias {day_start}-{day_end}")
                    return None, None
            else:
                # Carregar todos os períodos do mês
                cursor = conn.execute("""
                    SELECT prices_data, updated_at, day_start, day_end
                    FROM current_prices 
                    WHERE location = ? AND month = ? AND year = ?
                    ORDER BY day_start
                """, (location, month, year))
                
                rows = cursor.fetchall()
                
                periods = []
                for row in rows:
                    # Converter datetime para string se necessário (SQLite retorna string, mas por segurança)
                    updated_at_str = None
                    if row[1]:
                        if hasattr(row[1], 'isoformat'):
                            updated_at_str = row[1].isoformat()
                        else:
                            updated_at_str = str(row[1])
                    
                    periods.append({
                        'prices': json.loads(row[0]),
                        'updated_at': updated_at_str,
                        'day_start': row[2] if row[2] is not None else 1,
                        'day_end': row[3] if row[3] is not None else 31
                    })
                
                if periods:
                    logging.info(f"✅ {len(periods)} período(s) carregado(s): {location}, mês {month}/{year}")
                else:
                    logging.info(f"ℹ️ Sem preços para: {location}, mês {month}/{year}")
                
                return periods, None
                
    except Exception as e:
        logging.error(f"Erro ao carregar preços: {e}")
        return None, None

def save_prices_to_db(conn, location, month, year, prices_data, day_start=1, day_end=31):
    """Guarda preços na base de dados para um período específico"""
    try:
        prices_json = json.dumps(prices_data)
        
        # Detectar se é PostgreSQL ou SQLite
        is_postgres = conn.__class__.__module__ in ['psycopg2.extensions', 'psycopg2._psycopg']
        
        if is_postgres:
            # PostgreSQL
            with conn.cursor() as cur:
                # Verificar se já existe este período
                cur.execute("""
                    SELECT id FROM current_prices 
                    WHERE location = %s AND month = %s AND year = %s 
                      AND day_start = %s AND day_end = %s
                """, (location, month, year, day_start, day_end))
                
                existing = cur.fetchone()
                
                if existing:
                    # Atualizar
                    cur.execute("""
                        UPDATE current_prices 
                        SET prices_data = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE location = %s AND month = %s AND year = %s 
                          AND day_start = %s AND day_end = %s
                    """, (prices_json, location, month, year, day_start, day_end))
                else:
                    # Inserir
                    cur.execute("""
                        INSERT INTO current_prices (location, month, year, day_start, day_end, prices_data, updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                    """, (location, month, year, day_start, day_end, prices_json))
            conn.commit()
        else:
            # SQLite
            cursor = conn.execute("""
                SELECT id FROM current_prices 
                WHERE location = ? AND month = ? AND year = ? 
                  AND day_start = ? AND day_end = ?
            """, (location, month, year, day_start, day_end))
            
            existing = cursor.fetchone()
            
            if existing:
                # Atualizar
                conn.execute("""
                    UPDATE current_prices 
                    SET prices_data = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE location = ? AND month = ? AND year = ? 
                      AND day_start = ? AND day_end = ?
                """, (prices_json, location, month, year, day_start, day_end))
            else:
                # Inserir
                conn.execute("""
                    INSERT INTO current_prices (location, month, year, day_start, day_end, prices_data, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (location, month, year, day_start, day_end, prices_json))
            
            conn.commit()
        
        logging.info(f"✅ Preços guardados: {location}, mês {month}/{year}, dias {day_start}-{day_end}")
        return True
    except Exception as e:
        logging.error(f"Erro ao guardar preços: {e}")
        return False

def import_from_brokers_excel(file_path):
    """Importa preços de um ficheiro Excel Brokers"""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb['PREÇOS']
        
        prices = {}
        
        # Mapear colunas para dias
        # Row 4 tem os dias: coluna 2=1dia, 5=2dias, 8=3dias, etc
        day_columns = {
            1: 3,   # Coluna C
            2: 6,   # Coluna F
            3: 9,   # Coluna I
            4: 12,  # Coluna L
            5: 15,  # Coluna O
            6: 18,  # Coluna R
            7: 21,  # Coluna U
            8: 24,  # Coluna X
            9: 27,  # Coluna AA
            14: 30, # Coluna AD
            22: 33, # Coluna AG
            28: 36, # Coluna AJ
            31: 39, # Coluna AM
            60: 42  # Coluna AP
        }
        
        # Ler preços (começam na linha 5)
        for row_idx in range(5, 25):  # Linhas 5-24 (20 grupos)
            grupo = ws.cell(row_idx, 1).value
            if not grupo or grupo not in GRUPOS:
                continue
            
            prices[grupo] = {}
            
            for dia, col_idx in day_columns.items():
                net_value = ws.cell(row_idx, col_idx).value
                
                # Se for fórmula, calcular valor
                if isinstance(net_value, str) and net_value.startswith('='):
                    net_value = 0
                elif net_value is None:
                    net_value = 0
                else:
                    net_value = float(net_value)
                
                prices[grupo][dia] = {
                    'net': net_value,
                    'commission': calculate_commission(net_value)
                }
        
        return prices
    except Exception as e:
        logging.error(f"Erro ao importar Excel Brokers: {e}")
        raise

def generate_abbycar_excel(location, month, year, prices_data):
    """Gera ficheiro Excel no formato Abbycar"""
    try:
        # Carregar template Abbycar
        template_path = os.path.join(os.path.dirname(__file__), 'Abbycar.csv')
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Preços"
        
        # Cabeçalho
        month_names = ['', 'JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
        ws['A1'] = f"ABBYCAR - {location.upper()} - {month_names[month]} {year}"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['SIPP', 'Grupo'] + [f'{d}d' for d in DIAS]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(2, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Dados
        row_idx = 3
        for grupo in GRUPOS:
            ws.cell(row_idx, 1, f"{grupo}-CAR")
            ws.cell(row_idx, 2, grupo)
            
            # Verificar se é grupo especial
            is_group_k = grupo in GRUPOS_K
            is_commercial = grupo in GRUPOS_COMERCIAIS
            is_group_x = grupo == 'X'
            
            for col_idx, dia in enumerate(DIAS, 3):
                price_data = prices_data.get(grupo, {}).get(str(dia), {})
                net = price_data.get('net', 0) if price_data else 0
                comm = price_data.get('commission', 0) if price_data else 0
                
                # REGRA: Grupo X sempre sem valores
                if is_group_x:
                    price = 0
                # REGRA: C3,C4,C5 sempre NET (GROSS = NET)
                elif is_commercial and net > 0:
                    price = net
                # REGRA: Grupos K sem valores se não tiverem preços
                else:
                    price = comm if comm > 0 else 0
                
                ws.cell(row_idx, col_idx, price)
            
            row_idx += 1
        
        # Salvar em BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"ABBYCAR-{location.upper()}-{month_names[month]}-{year}.xlsx"
        return excel_file, filename
        
    except Exception as e:
        logging.error(f"Erro ao gerar Excel Abbycar: {e}")
        raise

def generate_brokers_excel(location, month, year, prices_data):
    """Gera ficheiro Excel no formato Brokers"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PREÇOS"
        
        month_names = ['', 'JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 
                       'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
        
        # Título
        ws['A1'] = month_names[month]
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Cabeçalho de dias (row 4)
        ws['A4'] = 'GRUPOS'
        col_idx = 2
        for dia in DIAS:
            ws.cell(4, col_idx, dia)
            ws.cell(4, col_idx).font = Font(bold=True)
            ws.cell(4, col_idx).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx+1)
            col_idx += 3  # Pular 3 colunas (net, comissão, diferença)
        
        # Sub-cabeçalhos
        col_idx = 2
        for _ in DIAS:
            ws.cell(5, col_idx, 'Net')
            ws.cell(5, col_idx).font = Font(size=9, color="EA580C")
            ws.cell(5, col_idx+1, 'Com Comissão')
            ws.cell(5, col_idx+1).font = Font(size=9, color="059669")
            ws.cell(5, col_idx+2, 'Dif %')
            ws.cell(5, col_idx+2).font = Font(size=9, color="6B7280")
            col_idx += 3
        
        # Dados dos grupos
        row_idx = 6
        for grupo in GRUPOS:
            ws.cell(row_idx, 1, grupo)
            ws.cell(row_idx, 1).font = Font(bold=True)
            
            # Verificar se é grupo especial
            is_group_k = grupo in GRUPOS_K
            is_commercial = grupo in GRUPOS_COMERCIAIS
            is_group_x = grupo == 'X'
            
            col_idx = 2
            for dia in DIAS:
                price_data = prices_data.get(grupo, {}).get(str(dia), {'net': 0, 'commission': 0})
                net = price_data.get('net', 0) if price_data else 0
                comm = price_data.get('commission', 0) if price_data else 0
                
                # REGRA: Grupo X sempre sem valores
                if is_group_x:
                    net = 0
                    comm = 0
                
                # REGRA: C3,C4,C5 sempre NET (GROSS = NET)
                if is_commercial and net > 0:
                    comm = net
                
                # REGRA: Grupos K sem valores se não tiverem preços (checkbox não selecionada)
                # Se não há preços no prices_data, deixar 0
                
                # Net
                ws.cell(row_idx, col_idx, net)
                ws.cell(row_idx, col_idx).number_format = '0.00'
                ws.cell(row_idx, col_idx).fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
                
                # Com Comissão (GROSS)
                ws.cell(row_idx, col_idx+1, comm)
                ws.cell(row_idx, col_idx+1).number_format = '0.00'
                ws.cell(row_idx, col_idx+1).fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
                
                # Diferença %
                if is_commercial:
                    diff_pct = 0  # Comerciais não têm comissão
                else:
                    diff_pct = COMMISSION_RATE * 100
                ws.cell(row_idx, col_idx+2, diff_pct)
                ws.cell(row_idx, col_idx+2).number_format = '0.00"%"'
                
                col_idx += 3
            
            row_idx += 1
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 12
        for col in range(2, col_idx):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 11
        
        # Salvar
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"Brokers-{location}-{month_names[month]}-{year}.xlsx"
        return excel_file, filename
        
    except Exception as e:
        logging.error(f"Erro ao gerar Excel Brokers: {e}")
        raise

def generate_website_excel(location, month, year, prices_data):
    """Gera ficheiro Excel no formato Website"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Preços Website"
        
        month_names = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
                       'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        
        # Título
        ws['A1'] = f"Preços Website - {location} - {month_names[month]} {year}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # Cabeçalhos
        headers = ['Grupo', 'Categoria'] + [f'{d} dia{"s" if d > 1 else ""}' for d in DIAS]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(3, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Mapear grupos para categorias
        group_categories = {
            'B1': 'Mini 4 Lugares', 'B2': 'Mini 5 Lugares',
            'BK1': 'Mini 4 Lugares Auto', 'BK2': 'Mini 5 Lugares Auto',
            'C': 'Economy', 'D': 'Economy Plus',
            'DK': 'Economy Auto',
            'E1': 'Mini Auto', 'E2': 'Economy Auto',
            'EK1': 'Compact Auto', 'EK2': 'Compact Plus Auto',
            'F': 'SUV', 'FK': 'SUV Auto',
            'G': 'Premium', 'J1': 'Crossover', 'J2': 'Station Wagon',
            'L1': '7 Lugares', 'L2': '7 Lugares Auto',
            'M1': '9 Lugares', 'M2': '9 Lugares Auto'
        }
        
        # Dados
        row_idx = 4
        for grupo in GRUPOS:
            ws.cell(row_idx, 1, grupo)
            ws.cell(row_idx, 1).font = Font(bold=True)
            ws.cell(row_idx, 2, group_categories.get(grupo, grupo))
            
            # Verificar se é grupo especial
            is_group_k = grupo in GRUPOS_K
            is_commercial = grupo in GRUPOS_COMERCIAIS
            is_group_x = grupo == 'X'
            
            for col_idx, dia in enumerate(DIAS, 3):
                price_data = prices_data.get(grupo, {}).get(str(dia), {})
                net = price_data.get('net', 0) if price_data else 0
                comm = price_data.get('commission', 0) if price_data else 0
                
                # REGRA: Grupo X sempre sem valores
                if is_group_x:
                    price = 0
                # REGRA: C3,C4,C5 sempre NET (GROSS = NET)
                elif is_commercial and net > 0:
                    price = net
                # REGRA: Grupos K sem valores se não tiverem preços
                else:
                    price = comm if comm > 0 else 0
                
                ws.cell(row_idx, col_idx, price)
                ws.cell(row_idx, col_idx).number_format = '0.00"€"'
                ws.cell(row_idx, col_idx).alignment = Alignment(horizontal='center')
            
            row_idx += 1
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        for col_idx in range(3, 3 + len(DIAS)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
        
        # Salvar
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"Website-{location}-{month_names[month]}-{year}.xlsx"
        return excel_file, filename
        
    except Exception as e:
        logging.error(f"Erro ao gerar Excel Website: {e}")
        raise

def create_current_prices_table(conn):
    """Cria tabela current_prices se não existir"""
    try:
        # Detectar se é PostgreSQL ou SQLite
        is_postgres = conn.__class__.__module__ in ['psycopg2.extensions', 'psycopg2._psycopg']
        
        if is_postgres:
            # PostgreSQL syntax
            with conn.cursor() as cur:
                # Criar tabela se não existir (sem UNIQUE constraint primeiro)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS current_prices (
                        id SERIAL PRIMARY KEY,
                        location TEXT NOT NULL,
                        month INTEGER NOT NULL,
                        year INTEGER NOT NULL,
                        prices_data TEXT NOT NULL,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Verificar se colunas day_start e day_end existem
                cur.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'current_prices' 
                    AND column_name IN ('day_start', 'day_end')
                """)
                existing_cols = [row[0] for row in cur.fetchall()]
                
                # Adicionar colunas se não existirem
                if 'day_start' not in existing_cols:
                    try:
                        cur.execute("ALTER TABLE current_prices ADD COLUMN day_start INTEGER DEFAULT 1")
                        conn.commit()
                        logging.info("Coluna day_start adicionada")
                    except Exception as e:
                        conn.rollback()  # Rollback em caso de erro
                        if "already exists" not in str(e).lower():
                            logging.error(f"Erro ao adicionar day_start: {e}")
                
                if 'day_end' not in existing_cols:
                    try:
                        cur.execute("ALTER TABLE current_prices ADD COLUMN day_end INTEGER DEFAULT 31")
                        conn.commit()
                        logging.info("Coluna day_end adicionada")
                    except Exception as e:
                        conn.rollback()  # Rollback em caso de erro
                        if "already exists" not in str(e).lower():
                            logging.error(f"Erro ao adicionar day_end: {e}")
                
                # Atualizar registos antigos sem day_start/day_end
                cur.execute("""
                    UPDATE current_prices 
                    SET day_start = COALESCE(day_start, 1), 
                        day_end = COALESCE(day_end, 31)
                    WHERE day_start IS NULL OR day_end IS NULL
                """)
                updated_rows = cur.rowcount
                if updated_rows > 0:
                    logging.info(f"Atualizados {updated_rows} registos com day_start/day_end padrão")
                
                # Verificar se constraint antiga existe
                cur.execute("""
                    SELECT constraint_name 
                    FROM information_schema.table_constraints 
                    WHERE table_name = 'current_prices' 
                    AND constraint_name = 'current_prices_location_month_year_key'
                """)
                old_constraint = cur.fetchone()
                
                if old_constraint:
                    cur.execute("""
                        ALTER TABLE current_prices 
                        DROP CONSTRAINT current_prices_location_month_year_key
                    """)
                    logging.info("Constraint antiga removida")
                
                # Verificar se nova constraint existe
                cur.execute("""
                    SELECT constraint_name 
                    FROM information_schema.table_constraints 
                    WHERE table_name = 'current_prices' 
                    AND constraint_name = 'current_prices_location_month_year_period_key'
                """)
                new_constraint = cur.fetchone()
                
                if not new_constraint:
                    cur.execute("""
                        ALTER TABLE current_prices 
                        ADD CONSTRAINT current_prices_location_month_year_period_key 
                        UNIQUE(location, month, year, day_start, day_end)
                    """)
                    logging.info("Nova constraint única adicionada")
                    
            conn.commit()
            logging.info("Tabela current_prices criada/migrada com sucesso")
        else:
            # SQLite syntax
            conn.execute("""
                CREATE TABLE IF NOT EXISTS current_prices (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    location TEXT NOT NULL,
                    month INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    day_start INTEGER DEFAULT 1,
                    day_end INTEGER DEFAULT 31,
                    prices_data TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(location, month, year, day_start, day_end)
                )
            """)
            # Adicionar colunas se não existirem (migração)
            try:
                conn.execute("ALTER TABLE current_prices ADD COLUMN day_start INTEGER DEFAULT 1")
            except:
                pass
            try:
                conn.execute("ALTER TABLE current_prices ADD COLUMN day_end INTEGER DEFAULT 31")
            except:
                pass
            conn.commit()
        
        logging.info("Tabela current_prices criada/verificada com sucesso")
    except Exception as e:
        logging.error(f"Erro ao criar tabela current_prices: {e}")
