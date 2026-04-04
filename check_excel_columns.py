import openpyxl

file_path = '/Users/filipepacheco/CascadeProjects/carscraping/CM-03-2026.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

print("=" * 80)
print("ESTRUTURA DO FICHEIRO CM-03-2026.xlsx")
print("=" * 80)

# Mostrar cabeçalhos (primeira linha)
print("\nCABEÇALHOS:")
for col in range(1, 10):
    header = ws.cell(row=1, column=col).value
    print(f"Coluna {col}: {header}")

# Mostrar exemplo de dados (linha 2)
print("\nEXEMPLO DE DADOS (Linha 2):")
for col in range(1, 10):
    value = ws.cell(row=2, column=col).value
    print(f"Coluna {col}: {value}")

# Mostrar linha com broker
print("\nEXEMPLO DE BROKER (Linha 3):")
for col in range(1, 10):
    value = ws.cell(row=3, column=col).value
    print(f"Coluna {col}: {value}")

# Mostrar linha com reserva
print("\nEXEMPLO DE RESERVA (Linha 4):")
for col in range(1, 10):
    value = ws.cell(row=4, column=col).value
    print(f"Coluna {col}: {value}")
