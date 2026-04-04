import psycopg2
import openpyxl
from datetime import datetime
import os

DATABASE_URL = "postgresql://postgres:OMxLodDbSGnDJUQGVIkXSXYxfiRwQqFo@shortline.proxy.rlwy.net:45408/railway"

def normalize_price(value):
    """Normalizar valores de preço"""
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip().replace('€', '').replace(' ', '')
    value_str = value_str.replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0

def parse_date(date_value):
    """Converter valores de data para date object"""
    if date_value is None or date_value == '':
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    if hasattr(date_value, 'date'):
        return date_value.date()
    date_str = str(date_value).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(date_str, fmt).date()
        except:
            continue
    return None

# Lista de nomes de brokers conhecidos
BROKER_NAMES = [
    'ABBYCAR-POA', 'ABBYCAR-PREPAID', 'AP', 'API', 'API-WEB',
    'CARJET-PREPAID', 'CARJET-POA', 'CARALLIANCE-PREPAID', 'CARALLIANCE-POA',
    'DISCOVERCARS-POA', 'DISCOVERCARS-PREPAID',
    'RENTALCARS', 'VIP CARS-POA', 'VIP CARS', 'BROKERS - DIRECTOS'
]

try:
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    
    print("\n" + "="*70)
    print("REIMPORTAÇÃO COMPLETA DE BROKERS 2025 E 2026")
    print("="*70)
    
    # 1. APAGAR TODOS OS DADOS DE BROKERS
    print("\n[1/3] A apagar todos os dados de brokers...")
    cur.execute("DELETE FROM broker_bookings")
    deleted = cur.rowcount
    conn.commit()
    print(f"✓ {deleted} registos apagados")
    
    # 2. IMPORTAR 2025
    print("\n[2/3] A importar brokers de 2025...")
    excel_dir_2025 = "./CM-25"
    files_2025 = sorted([f for f in os.listdir(excel_dir_2025) if f.startswith('CM-') and f.endswith('.xlsx') and not f.startswith('~$')])
    
    total_2025 = 0
    skipped_2025 = 0
    
    for filename in files_2025:
        filepath = os.path.join(excel_dir_2025, filename)
        print(f"\n  Ficheiro: {filename}")
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        current_broker = None
        file_imported = 0
        file_skipped = 0
        
        for row_idx in range(1, ws.max_row + 1):
            col_a = ws.cell(row_idx, 1).value  # Voucher ou nome do broker
            col_b = ws.cell(row_idx, 2).value  # Data Entrega
            col_c = ws.cell(row_idx, 3).value  # Dias
            col_d = ws.cell(row_idx, 4).value  # Loyalty Card (preço base)
            
            # Ignorar cabeçalho
            if col_a and 'VOUCHER' in str(col_a).upper():
                continue
            
            # Verificar se é linha de broker (nome na col A, sem data na col B)
            if col_a and not col_b:
                broker_name = str(col_a).strip()
                if broker_name in BROKER_NAMES:
                    current_broker = broker_name
                continue
            
            # Se temos broker atual e linha tem data = é reserva
            if current_broker and col_b:
                # IMPORTANTE: Importar voucher original ou deixar NULL
                voucher = None
                if col_a and str(col_a).strip():
                    voucher_str = str(col_a).strip()
                    # Não validar se é número - importar qualquer voucher que exista
                    if voucher_str and voucher_str.upper() not in BROKER_NAMES:
                        voucher = voucher_str
                
                pickup_date = parse_date(col_b)
                # Converter dias - aceitar int, float ou string numérica
                days = None
                if col_c is not None and col_c != '':
                    try:
                        days = int(float(col_c))
                    except:
                        days = None
                total_price = normalize_price(col_d)
                
                if not pickup_date or total_price <= 0:
                    file_skipped += 1
                    continue
                
                # Inserir com voucher original (ou NULL se não existir)
                try:
                    cur.execute("""
                        INSERT INTO broker_bookings (
                            voucher_number, broker_name, pickup_date, days, 
                            total_price, status, created_at
                        ) VALUES (
                            %s, %s, %s, %s, %s, 'confirmed', NOW()
                        )
                    """, (
                        voucher,  # Voucher original ou NULL
                        current_broker,
                        pickup_date,
                        days,
                        total_price
                    ))
                    file_imported += 1
                except Exception as e:
                    print(f"    ✗ Erro linha {row_idx}: {e}")
                    file_skipped += 1
        
        conn.commit()
        total_2025 += file_imported
        skipped_2025 += file_skipped
        print(f"    ✓ Importados: {file_imported}, Ignorados: {file_skipped}")
        wb.close()
    
    print(f"\n  TOTAL 2025: {total_2025} importados, {skipped_2025} ignorados")
    
    # 3. IMPORTAR 2026
    print("\n[3/3] A importar brokers de 2026...")
    excel_dir_2026 = "./CM-26"
    files_2026 = sorted([f for f in os.listdir(excel_dir_2026) if f.startswith('CM-') and f.endswith('.xlsx') and not f.startswith('~$')])
    
    total_2026 = 0
    skipped_2026 = 0
    
    for filename in files_2026:
        filepath = os.path.join(excel_dir_2026, filename)
        print(f"\n  Ficheiro: {filename}")
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        current_broker = None
        file_imported = 0
        file_skipped = 0
        
        for row_idx in range(1, ws.max_row + 1):
            col_a = ws.cell(row_idx, 1).value  # Voucher ou nome do broker
            col_b = ws.cell(row_idx, 2).value  # Data Entrega
            col_c = ws.cell(row_idx, 3).value  # Dias
            col_d = ws.cell(row_idx, 4).value  # Loyalty Card (preço base)
            
            # Ignorar cabeçalho
            if col_a and 'VOUCHER' in str(col_a).upper():
                continue
            
            # Verificar se é linha de broker (nome na col A, sem data na col B)
            if col_a and not col_b:
                broker_name = str(col_a).strip()
                if broker_name in BROKER_NAMES:
                    current_broker = broker_name
                continue
            
            # Se temos broker atual e linha tem data = é reserva
            if current_broker and col_b:
                # IMPORTANTE: Importar voucher original ou deixar NULL
                voucher = None
                if col_a and str(col_a).strip():
                    voucher_str = str(col_a).strip()
                    # Não validar se é número - importar qualquer voucher que exista
                    if voucher_str and voucher_str.upper() not in BROKER_NAMES:
                        voucher = voucher_str
                
                pickup_date = parse_date(col_b)
                # Converter dias - aceitar int, float ou string numérica
                days = None
                if col_c is not None and col_c != '':
                    try:
                        days = int(float(col_c))
                    except:
                        days = None
                total_price = normalize_price(col_d)
                
                if not pickup_date or total_price <= 0:
                    file_skipped += 1
                    continue
                
                # Inserir com voucher original (ou NULL se não existir)
                try:
                    cur.execute("""
                        INSERT INTO broker_bookings (
                            voucher_number, broker_name, pickup_date, days, 
                            total_price, status, created_at
                        ) VALUES (
                            %s, %s, %s, %s, %s, 'confirmed', NOW()
                        )
                    """, (
                        voucher,  # Voucher original ou NULL
                        current_broker,
                        pickup_date,
                        days,
                        total_price
                    ))
                    file_imported += 1
                except Exception as e:
                    print(f"    ✗ Erro linha {row_idx}: {e}")
                    file_skipped += 1
        
        conn.commit()
        total_2026 += file_imported
        skipped_2026 += file_skipped
        print(f"    ✓ Importados: {file_imported}, Ignorados: {file_skipped}")
        wb.close()
    
    print(f"\n  TOTAL 2026: {total_2026} importados, {skipped_2026} ignorados")
    
    # RESUMO FINAL
    print("\n" + "="*70)
    print("RESUMO FINAL")
    print("="*70)
    print(f"2025: {total_2025} importados, {skipped_2025} ignorados")
    print(f"2026: {total_2026} importados, {skipped_2026} ignorados")
    print(f"TOTAL: {total_2025 + total_2026} importados, {skipped_2025 + skipped_2026} ignorados")
    print("\n✓ Reimportação concluída com sucesso!")
    print("  - Vouchers originais importados (ou NULL se não existir)")
    print("  - Dias importados corretamente")
    print("  - Loyalty Card (preço base) importado")
    print("  - Data de entrega importada")
    
    cur.close()
    conn.close()
    
except Exception as e:
    print(f"\n✗ Erro: {e}")
    import traceback
    traceback.print_exc()
