import openpyxl
from datetime import datetime

def normalize_price(value):
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip().replace('€', '').replace(' ', '')
    value_str = value_str.replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0

def parse_date(date_value):
    if date_value is None or date_value == '':
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    return None

BROKER_NAMES = [
    'ABBYCAR-POA', 'ABBYCAR-PREPAID', 'AP', 'API', 'API-WEB',
    'CARJET-PREPAID', 'CARJET-POA', 'DISCOVERCARS-POA', 'DISCOVERCARS-PREPAID',
    'RENTALCARS', 'VIP CARS-POA', 'VIP CARS', 'BROKERS - DIRECTOS'
]

# Analisar Janeiro
filepath = "./CM-25/CM-01-2025.xlsx"
print(f"\n=== ANÁLISE DETALHADA: {filepath} ===\n")

wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb.active

current_broker = None
total_lines = 0
valid_lines = 0
ignored_lines = 0
ignored_reasons = []

for row_idx in range(1, ws.max_row + 1):
    col_a = ws.cell(row_idx, 1).value
    col_b = ws.cell(row_idx, 2).value
    col_c = ws.cell(row_idx, 3).value
    col_d = ws.cell(row_idx, 4).value
    
    # Ignorar cabeçalho
    if col_a and 'VOUCHER' in str(col_a).upper():
        continue
    
    # Linha de broker
    if col_a and not col_b:
        broker_name = str(col_a).strip()
        if broker_name in BROKER_NAMES:
            current_broker = broker_name
            print(f"Linha {row_idx}: BROKER = {current_broker}")
        continue
    
    # Linha de dados
    if current_broker and col_b:
        total_lines += 1
        
        pickup_date = parse_date(col_b)
        price = normalize_price(col_d)
        
        reason = None
        if not pickup_date:
            reason = f"Data inválida: {col_b}"
        elif price <= 0:
            reason = f"Preço inválido: {col_d} (normalizado: {price})"
        
        if reason:
            ignored_lines += 1
            ignored_reasons.append(f"Linha {row_idx} ({current_broker}): {reason}")
            print(f"  ✗ IGNORADO Linha {row_idx}: {reason}")
            print(f"    Col A: {col_a}, Col B: {col_b}, Col C: {col_c}, Col D: {col_d}")
        else:
            valid_lines += 1

wb.close()

print(f"\n=== RESUMO ===")
print(f"Total de linhas de dados: {total_lines}")
print(f"Linhas válidas: {valid_lines}")
print(f"Linhas ignoradas: {ignored_lines}")

if ignored_reasons:
    print(f"\n=== RAZÕES DE IGNORAR ===")
    for reason in ignored_reasons:
        print(f"  {reason}")
