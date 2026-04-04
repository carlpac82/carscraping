import openpyxl
import os
from datetime import datetime

def normalize_price(value):
    """Normalizar valores de preço"""
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip().replace('€', '').replace(' ', '')
    value_str = value_str.replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0

def parse_date(date_value):
    """Converter valores de data para date object"""
    if date_value is None or date_value == '':
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    if hasattr(date_value, 'date'):
        return date_value.date()
    date_str = str(date_value).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(date_str, fmt).date()
        except:
            continue
    return None

BROKER_NAMES = [
    'ABBYCAR-POA', 'ABBYCAR-PREPAID', 'AP', 'API', 'API-WEB',
    'CARJET-PREPAID', 'CARJET-POA', 'CARALLIANCE-PREPAID', 'CARALLIANCE-POA',
    'DISCOVERCARS-POA', 'DISCOVERCARS-PREPAID',
    'RENTALCARS', 'VIP CARS-POA', 'VIP CARS', 'BROKERS - DIRECTOS'
]

print("\n" + "="*70)
print("ANÁLISE DE REGISTOS IGNORADOS EM 2025")
print("="*70)

excel_dir = "./CM-25"
files = sorted([f for f in os.listdir(excel_dir) if f.startswith('CM-') and f.endswith('.xlsx') and not f.startswith('~$')])

total_ignored = 0
ignored_details = []

for filename in files:
    filepath = os.path.join(excel_dir, filename)
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    current_broker = None
    file_ignored = 0
    
    for row_idx in range(1, ws.max_row + 1):
        col_a = ws.cell(row_idx, 1).value  # Voucher ou nome do broker
        col_b = ws.cell(row_idx, 2).value  # Data Entrega
        col_c = ws.cell(row_idx, 3).value  # Dias
        col_d = ws.cell(row_idx, 4).value  # Loyalty Card (preço base)
        
        # Ignorar cabeçalho
        if col_a and 'VOUCHER' in str(col_a).upper():
            continue
        
        # Verificar se é linha de broker
        if col_a and not col_b:
            broker_name = str(col_a).strip()
            if broker_name in BROKER_NAMES:
                current_broker = broker_name
            continue
        
        # Se temos broker atual e linha tem data = é reserva
        if current_broker and col_b:
            pickup_date = parse_date(col_b)
            total_price = normalize_price(col_d)
            
            # Verificar se seria ignorado
            if not pickup_date or total_price <= 0:
                file_ignored += 1
                total_ignored += 1
                
                reason = []
                if not pickup_date:
                    reason.append("Data inválida")
                if total_price <= 0:
                    reason.append(f"Preço inválido ({col_d})")
                
                ignored_details.append({
                    'file': filename,
                    'row': row_idx,
                    'broker': current_broker,
                    'voucher': col_a,
                    'date': col_b,
                    'days': col_c,
                    'price': col_d,
                    'reason': ' | '.join(reason)
                })
    
    if file_ignored > 0:
        print(f"\n{filename}: {file_ignored} ignorados")
    
    wb.close()

print("\n" + "="*70)
print(f"TOTAL DE REGISTOS IGNORADOS: {total_ignored}")
print("="*70)

if ignored_details:
    print("\nDETALHES DOS REGISTOS IGNORADOS:\n")
    for detail in ignored_details:
        print(f"Ficheiro: {detail['file']}")
        print(f"  Linha: {detail['row']}")
        print(f"  Broker: {detail['broker']}")
        print(f"  Voucher: {detail['voucher']}")
        print(f"  Data: {detail['date']}")
        print(f"  Dias: {detail['days']}")
        print(f"  Preço: {detail['price']}")
        print(f"  Razão: {detail['reason']}")
        print()
