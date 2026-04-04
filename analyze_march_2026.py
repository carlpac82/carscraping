import openpyxl
from datetime import datetime
import psycopg2

DATABASE_URL = "postgresql://postgres:OMxLodDbSGnDJUQGVIkXSXYxfiRwQqFo@shortline.proxy.rlwy.net:45408/railway"

def analyze_march_2026_file():
    """Analisar o ficheiro CM-03-2026.xlsx"""
    
    file_path = '/Users/filipepacheco/CascadeProjects/carscraping/CM-03-2026.xlsx'
    
    print("=" * 80)
    print("ANÁLISE DO FICHEIRO CM-03-2026.xlsx")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"\n📄 Ficheiro: {file_path}")
        print(f"📊 Total de linhas: {ws.max_row}")
        
        # Analisar todas as linhas
        print("\n" + "=" * 80)
        print("TODAS AS LINHAS DO FICHEIRO:")
        print("=" * 80)
        print(f"\n{'Linha':<6} {'Col A (Voucher)':<20} {'Col B (Data)':<15} {'Col C (Dias)':<10} {'Col D (Price)':<10}")
        print("-" * 80)
        
        total_price = 0
        valid_bookings = 0
        broker_name = None
        
        for row_num in range(1, ws.max_row + 1):
            col_a = ws.cell(row=row_num, column=1).value
            col_b = ws.cell(row=row_num, column=2).value
            col_c = ws.cell(row=row_num, column=3).value
            col_d = ws.cell(row=row_num, column=4).value
            
            # Converter valores para string para display
            col_a_str = str(col_a) if col_a else ''
            col_b_str = str(col_b) if col_b else ''
            col_c_str = str(col_c) if col_c else ''
            col_d_str = str(col_d) if col_d else ''
            
            print(f"{row_num:<6} {col_a_str:<20} {col_b_str:<15} {col_c_str:<10} {col_d_str:<10}")
            
            # Verificar se é linha de broker (tem texto em A mas não tem data em B)
            if col_a and not col_b:
                broker_name = str(col_a).strip()
                print(f"       ↑ BROKER DETECTADO: {broker_name}")
            
            # Verificar se é linha de booking (tem data em B)
            elif col_b:
                # Tentar obter o preço
                price = 0
                if col_d:
                    try:
                        price = float(col_d)
                        total_price += price
                        valid_bookings += 1
                    except:
                        pass
        
        print("\n" + "=" * 80)
        print("RESUMO DO FICHEIRO:")
        print("=" * 80)
        print(f"Total de reservas válidas: {valid_bookings}")
        print(f"Soma total dos preços: €{total_price:.2f}")
        
        # Verificar o que está na base de dados
        print("\n" + "=" * 80)
        print("DADOS NA BASE DE DADOS PARA MARÇO 2026:")
        print("=" * 80)
        
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        
        query = """
            SELECT 
                COUNT(*) as count,
                SUM(price) as total_price,
                SUM(commission_amount) as total_commission
            FROM commission_bookings
            WHERE EXTRACT(YEAR FROM pickup_date) = 2026
            AND EXTRACT(MONTH FROM pickup_date) = 3
        """
        
        cur.execute(query)
        row = cur.fetchone()
        
        db_count = row[0]
        db_price = float(row[1]) if row[1] else 0
        db_commission = float(row[2]) if row[2] else 0
        
        print(f"\nReservas na BD: {db_count}")
        print(f"Total Price na BD: €{db_price:.2f}")
        print(f"Total Comissão na BD: €{db_commission:.2f}")
        
        # Mostrar alguns exemplos
        print("\n" + "=" * 80)
        print("EXEMPLOS DE REGISTOS NA BD (Março 2026):")
        print("=" * 80)
        
        query = """
            SELECT 
                cb.voucher_number,
                cb.pickup_date,
                cb.price,
                cb.commission_amount,
                c.name
            FROM commission_bookings cb
            JOIN commissioners c ON cb.commissioner_id = c.id
            WHERE EXTRACT(YEAR FROM cb.pickup_date) = 2026
            AND EXTRACT(MONTH FROM cb.pickup_date) = 3
            ORDER BY cb.pickup_date
            LIMIT 20
        """
        
        cur.execute(query)
        rows = cur.fetchall()
        
        print(f"\n{'Voucher':<20} {'Data':<12} {'Price':<10} {'Comissão':<10} {'Comissionista':<20}")
        print("-" * 80)
        
        for row in rows:
            voucher = row[0] or 'N/A'
            date = row[1]
            price = float(row[2]) if row[2] else 0
            commission = float(row[3]) if row[3] else 0
            name = row[4]
            
            print(f"{voucher:<20} {date} €{price:<9.2f} €{commission:<9.2f} {name:<20}")
        
        cur.close()
        conn.close()
        
        # Comparação
        print("\n" + "=" * 80)
        print("COMPARAÇÃO:")
        print("=" * 80)
        print(f"Ficheiro Excel: {valid_bookings} reservas, €{total_price:.2f}")
        print(f"Base de Dados:  {db_count} reservas, €{db_price:.2f}")
        
        if abs(total_price - db_price) > 0.01:
            print(f"\n⚠️  DIFERENÇA DETECTADA: €{abs(total_price - db_price):.2f}")
        
    except Exception as e:
        print(f"\n❌ Erro ao analisar ficheiro: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_march_2026_file()
