import psycopg2
import openpyxl
from datetime import datetime
import os

DATABASE_URL = "postgresql://postgres:OMxLodDbSGnDJUQGVIkXSXYxfiRwQqFo@shortline.proxy.rlwy.net:45408/railway"

def normalize_price(value):
    """Normalizar valores de preço"""
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip().replace('€', '').replace(' ', '')
    value_str = value_str.replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0

def parse_date(date_value):
    """Parse date from various formats"""
    if date_value is None or date_value == '':
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    date_str = str(date_value).strip()
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(date_str, fmt).date()
        except:
            continue
    return None

def get_commissioners_map(conn):
    """Obter mapa de comissionistas: nome -> (id, commission_rate)"""
    cur = conn.cursor()
    cur.execute("SELECT id, name, commission_rate FROM commissioners")
    commissioners = {}
    for row in cur.fetchall():
        # Normalizar nome para uppercase e remover espaços extras
        name_normalized = row[1].strip().upper()
        commissioners[name_normalized] = (row[0], row[1])
    cur.close()
    return commissioners

def find_commissioner_id(broker_name, commissioners_map):
    """Encontrar ID do comissionista pelo nome"""
    if not broker_name:
        return None
    
    # Normalizar nome
    name_normalized = broker_name.strip().upper()
    
    # Procura exata
    if name_normalized in commissioners_map:
        return commissioners_map[name_normalized][0]
    
    # Procura parcial
    for comm_name, (comm_id, original_name) in commissioners_map.items():
        if name_normalized in comm_name or comm_name in name_normalized:
            return comm_id
    
    return None

try:
    conn = psycopg2.connect(DATABASE_URL)
    
    # Obter mapa de comissionistas
    print("A carregar comissionistas da base de dados...")
    commissioners_map = get_commissioners_map(conn)
    print(f"✓ {len(commissioners_map)} comissionistas carregados\n")
    
    # Limpar dados de 2025 em commission_bookings
    print("A limpar dados de 2025 em commission_bookings...")
    cur = conn.cursor()
    cur.execute("""
        DELETE FROM commission_bookings 
        WHERE EXTRACT(YEAR FROM pickup_date) = 2025
    """)
    deleted = cur.rowcount
    conn.commit()
    print(f"✓ {deleted} registos removidos\n")
    
    # Processar ficheiros Excel
    excel_dir = "./2025"
    files = sorted([f for f in os.listdir(excel_dir) if f.startswith('CM-') and f.endswith('.xlsx')])
    
    total_imported = 0
    total_skipped = 0
    not_found_commissioners = set()
    
    for filename in files:
        filepath = os.path.join(excel_dir, filename)
        print(f"A processar {filename}...")
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Encontrar linha de cabeçalho
        header_row = None
        for row_idx in range(1, 20):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and 'VOUCHER' in str(cell_value).upper():
                header_row = row_idx
                break
        
        if not header_row:
            print(f"  ✗ Cabeçalho não encontrado")
            continue
        
        # Processar linhas de dados
        current_broker = None
        file_imported = 0
        file_skipped = 0
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            voucher = ws.cell(row_idx, 1).value
            
            # Verificar se é linha de broker
            if voucher and isinstance(voucher, str) and not voucher.replace('-', '').isdigit():
                current_broker = voucher.strip()
                continue
            
            # Verificar se é linha de dados
            if not voucher or not str(voucher).strip():
                continue
            
            # Extrair dados
            client_name = ws.cell(row_idx, 2).value
            pickup_date = parse_date(ws.cell(row_idx, 3).value)
            dropoff_date = parse_date(ws.cell(row_idx, 4).value)
            price = normalize_price(ws.cell(row_idx, 5).value)
            
            if not pickup_date or price <= 0:
                file_skipped += 1
                continue
            
            # Encontrar commissioner_id
            commissioner_id = find_commissioner_id(current_broker, commissioners_map)
            
            if not commissioner_id:
                if current_broker:
                    not_found_commissioners.add(current_broker)
                file_skipped += 1
                continue
            
            # Calcular comissão (15% por defeito)
            commission_rate = 15.0
            commission_amount = price * (commission_rate / 100)
            
            # Inserir em commission_bookings
            try:
                cur.execute("""
                    INSERT INTO commission_bookings (
                        voucher_number, client_name, pickup_date, dropoff_date,
                        price, commissioner_id, commission_rate, commission_amount,
                        commission_paid, status, created_at, updated_at,
                        pickup_location, dropoff_location, vehicle_group
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, false, 'confirmed',
                        NOW(), NOW(), 'N/A', 'N/A', 'N/A'
                    )
                """, (
                    str(voucher).strip(),
                    client_name or 'N/A',
                    pickup_date,
                    dropoff_date,
                    price,
                    commissioner_id,
                    commission_rate,
                    commission_amount
                ))
                file_imported += 1
            except Exception as e:
                if 'unique constraint' in str(e).lower():
                    file_skipped += 1
                else:
                    print(f"  ✗ Erro ao inserir voucher {voucher}: {e}")
                    file_skipped += 1
        
        conn.commit()
        total_imported += file_imported
        total_skipped += file_skipped
        print(f"  ✓ Importados: {file_imported}, Ignorados: {file_skipped}")
    
    cur.close()
    conn.close()
    
    print(f"\n=== RESUMO ===")
    print(f"Total importado: {total_imported}")
    print(f"Total ignorado: {total_skipped}")
    
    if not_found_commissioners:
        print(f"\n⚠ Comissionistas não encontrados na BD ({len(not_found_commissioners)}):")
        for name in sorted(not_found_commissioners):
            print(f"  - {name}")
    
    print("\n✓ Importação concluída!")
    
except Exception as e:
    print(f"✗ Erro: {e}")
    import traceback
    traceback.print_exc()
