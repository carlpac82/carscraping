import openpyxl
import psycopg2
from datetime import datetime, timedelta
import os

DATABASE_URL = "postgresql://postgres:OMxLodDbSGnDJUQGVIkXSXYxfiRwQqFo@shortline.proxy.rlwy.net:45408/railway"

def parse_date(date_value):
    """Parse date from Excel"""
    if date_value is None or date_value == '':
        return None
    if isinstance(date_value, datetime):
        return date_value
    if hasattr(date_value, 'date'):
        return date_value
    date_str = str(date_value).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(date_str, fmt)
        except:
            continue
    return None

def get_commissioners_map(conn):
    """Obter mapa de comissionistas: nome -> (id, commission_rate)"""
    cur = conn.cursor()
    cur.execute("SELECT id, name, commission_rate FROM commissioners")
    commissioners = {}
    for row in cur.fetchall():
        name_normalized = row[1].strip().upper()
        commissioners[name_normalized] = (row[0], float(row[2]) if row[2] else 0.15)
    cur.close()
    return commissioners

def find_commissioner(broker_name, commissioners_map):
    """Encontrar ID e taxa do comissionista pelo nome"""
    if not broker_name:
        return None, None
    
    broker_normalized = broker_name.strip().upper()
    
    if broker_normalized in commissioners_map:
        return commissioners_map[broker_normalized]
    
    for comm_name, (comm_id, comm_rate) in commissioners_map.items():
        if comm_name in broker_normalized or broker_normalized in comm_name:
            return comm_id, comm_rate
    
    return None, None

def import_month(file_path, month_name, conn, commissioners_map):
    """Importar dados de um mês"""
    
    print(f"\n{'=' * 80}")
    print(f"IMPORTAR {month_name.upper()}")
    print(f"{'=' * 80}")
    
    if not os.path.exists(file_path):
        print(f"❌ Ficheiro não encontrado: {file_path}")
        return 0, 0
    
    # Abrir ficheiro Excel
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"📄 Ficheiro: {file_path}")
    print(f"📊 Total de linhas: {ws.max_row}")
    
    cur = conn.cursor()
    current_broker = None
    commissioner_id = None
    commission_rate = None
    
    imported_count = 0
    ignored_count = 0
    
    for row_num in range(1, ws.max_row + 1):
        col_a = ws.cell(row=row_num, column=1).value  # Voucher ou Broker
        col_b = ws.cell(row=row_num, column=2).value  # Data Entrega
        col_c = ws.cell(row=row_num, column=3).value  # Dias
        col_d = ws.cell(row=row_num, column=4).value  # Price (Loyalty Card)
        
        # Linha de cabeçalho - ignorar
        if row_num == 1:
            continue
        
        # Linha de broker: tem texto em A mas não tem data em B
        if col_a and not col_b:
            current_broker = str(col_a).strip()
            commissioner_id, commission_rate = find_commissioner(current_broker, commissioners_map)
            
            if commissioner_id:
                print(f"🏨 {current_broker}")
            else:
                print(f"⚠️  {current_broker} - NÃO ENCONTRADO")
            continue
        
        # Linha de booking: tem data em B
        if col_b and current_broker and commissioner_id:
            # Parse da data
            pickup_datetime = parse_date(col_b)
            if not pickup_datetime:
                ignored_count += 1
                continue
            
            # Parse do preço (coluna D - Loyalty Card)
            price = 0
            if col_d:
                try:
                    price = float(col_d)
                except:
                    ignored_count += 1
                    continue
            
            if price <= 0:
                ignored_count += 1
                continue
            
            # Parse dos dias
            days = 1
            if col_c:
                try:
                    days = int(col_c)
                except:
                    days = 1
            
            # Voucher (se existir na coluna A)
            voucher = None
            if col_a:
                voucher_str = str(col_a).strip()
                if voucher_str and voucher_str.upper() != current_broker.upper():
                    voucher = voucher_str
            
            # Calcular dropoff_date
            pickup_date = pickup_datetime.date()
            pickup_time = pickup_datetime.strftime('%H:%M')
            dropoff_date = pickup_date + timedelta(days=days)
            
            # Calcular comissão: price × commission_rate
            commission_amount = price * commission_rate
            
            # Inserir na BD
            try:
                insert_query = """
                    INSERT INTO commission_bookings (
                        commissioner_id, voucher_number, client_name, client_email, client_phone,
                        pickup_date, pickup_time, dropoff_date, dropoff_time,
                        pickup_location, dropoff_location, vehicle_group, extras,
                        price, commission_amount, status, created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                        CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
                    )
                """
                
                cur.execute(insert_query, (
                    commissioner_id, voucher, 'Commission Client', 'commission@rentalprices.pt', '',
                    pickup_date, pickup_time, dropoff_date, '00:00',
                    '', '', '', '[]',
                    price, commission_amount, 'confirmed'
                ))
                
                imported_count += 1
                
            except Exception as e:
                print(f"  ❌ Erro linha {row_num}: {str(e)}")
                ignored_count += 1
    
    conn.commit()
    cur.close()
    
    print(f"\n✅ {month_name}: {imported_count} importadas, {ignored_count} ignoradas")
    
    return imported_count, ignored_count

def reimport_all_2026():
    """Reimportar todos os meses de 2026"""
    
    print("=" * 80)
    print("REIMPORTAR TODOS OS DADOS DE 2026")
    print("=" * 80)
    
    conn = psycopg2.connect(DATABASE_URL)
    
    # Obter mapa de comissionistas
    commissioners_map = get_commissioners_map(conn)
    print(f"\n📋 Comissionistas na BD: {len(commissioners_map)}")
    
    months = [
        ('CM-01-2026.xlsx', 'Janeiro 2026'),
        ('CM-02-2026.xlsx', 'Fevereiro 2026'),
    ]
    
    total_imported = 0
    total_ignored = 0
    
    for file_name, month_name in months:
        file_path = f'/Users/filipepacheco/CascadeProjects/carscraping/{file_name}'
        imported, ignored = import_month(file_path, month_name, conn, commissioners_map)
        total_imported += imported
        total_ignored += ignored
    
    # Verificar totais finais
    print("\n" + "=" * 80)
    print("RESUMO FINAL:")
    print("=" * 80)
    print(f"✅ Total importado: {total_imported}")
    print(f"⚠️  Total ignorado: {total_ignored}")
    
    # Verificar dados na BD
    cur = conn.cursor()
    
    for month in [1, 2, 3]:
        month_names = ['Janeiro', 'Fevereiro', 'Março']
        
        query = """
            SELECT 
                COUNT(*) as count,
                SUM(price) as total_price,
                SUM(commission_amount) as total_commission
            FROM commission_bookings
            WHERE EXTRACT(YEAR FROM pickup_date) = 2026
            AND EXTRACT(MONTH FROM pickup_date) = %s
        """
        
        cur.execute(query, (month,))
        row = cur.fetchone()
        
        count = row[0]
        total_price = float(row[1]) if row[1] else 0
        total_commission = float(row[2]) if row[2] else 0
        
        print(f"\n{month_names[month-1]} 2026:")
        print(f"  Reservas: {count}")
        print(f"  Total Price: €{total_price:.2f}")
        print(f"  Total Comissão: €{total_commission:.2f}")
    
    cur.close()
    conn.close()

if __name__ == "__main__":
    reimport_all_2026()
