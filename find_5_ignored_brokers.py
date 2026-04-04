import openpyxl
import os
from datetime import datetime

def normalize_price(value):
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip().replace('€', '').replace(' ', '')
    value_str = value_str.replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0

def parse_date(date_value):
    if date_value is None or date_value == '':
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    if hasattr(date_value, 'date'):
        return date_value.date()
    date_str = str(date_value).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(date_str, fmt).date()
        except:
            continue
    return None

BROKER_NAMES = [
    'ABBYCAR-POA', 'ABBYCAR-PREPAID', 'AP', 'API', 'API-WEB',
    'CARJET-PREPAID', 'CARJET-POA', 'DISCOVERCARS-POA', 'DISCOVERCARS-PREPAID',
    'RENTALCARS', 'VIP CARS-POA', 'VIP CARS', 'BROKERS - DIRECTOS'
]

excel_dir = "./CM-25"
files = sorted([f for f in os.listdir(excel_dir) if f.startswith('CM-') and f.endswith('.xlsx') and not f.startswith('~$')])

print("\n=== PROCURANDO OS 5 REGISTOS IGNORADOS ===\n")

total_ignored = 0

for filename in files:
    filepath = os.path.join(excel_dir, filename)
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    current_broker = None
    
    for row_idx in range(1, ws.max_row + 1):
        col_a = ws.cell(row_idx, 1).value
        col_b = ws.cell(row_idx, 2).value
        col_c = ws.cell(row_idx, 3).value
        col_d = ws.cell(row_idx, 4).value
        
        # Ignorar cabeçalho
        if col_a and 'VOUCHER' in str(col_a).upper():
            continue
        
        # Linha de broker
        if col_a and not col_b:
            broker_name = str(col_a).strip()
            if broker_name in BROKER_NAMES:
                current_broker = broker_name
            continue
        
        # Linha de dados
        if current_broker and col_b:
            pickup_date = parse_date(col_b)
            price = normalize_price(col_d)
            
            if not pickup_date or price <= 0:
                total_ignored += 1
                print(f"IGNORADO #{total_ignored} - {filename}, Linha {row_idx}, Broker: {current_broker}")
                print(f"  Voucher: {col_a}")
                print(f"  Data: {col_b} (parsed: {pickup_date})")
                print(f"  Dias: {col_c}")
                print(f"  Preço: {col_d} (normalizado: {price})")
                print()
    
    wb.close()

print(f"Total de registos ignorados encontrados: {total_ignored}")
