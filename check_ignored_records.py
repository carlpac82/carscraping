import openpyxl

# Verificar os 2 registos ignorados
files_to_check = [
    ("./2025/CM-01-2025.xlsx", "Janeiro"),
    ("./2025/CM-03-2025.xlsx", "Março")
]

for filepath, month in files_to_check:
    print(f"\n=== {month} - {filepath} ===\n")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    print(f"Total de linhas: {ws.max_row}\n")
    print("Primeiras 20 linhas:")
    
    for row_idx in range(1, min(21, ws.max_row + 1)):
        col1 = ws.cell(row_idx, 1).value
        col2 = ws.cell(row_idx, 2).value
        col3 = ws.cell(row_idx, 3).value
        col4 = ws.cell(row_idx, 4).value
        
        print(f"Linha {row_idx:2d}: {str(col1)[:30] if col1 else '':<30} | {str(col2)[:20] if col2 else '':<20} | {str(col3)[:10] if col3 else '':<10} | {str(col4)[:10] if col4 else ''}")
    
    wb.close()
