import openpyxl
import os

excel_dir = "./CM-25"
files = sorted([f for f in os.listdir(excel_dir) if f.startswith('CM-') and f.endswith('.xlsx') and not f.startswith('~$')])

all_broker_names = set()

for filename in files:
    filepath = os.path.join(excel_dir, filename)
    print(f"\n=== {filename} ===")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    file_brokers = []
    
    for row_idx in range(1, min(50, ws.max_row + 1)):
        col1 = ws.cell(row_idx, 1).value
        col2 = ws.cell(row_idx, 2).value
        
        # Se tem valor na col1 mas não na col2, pode ser broker
        if col1 and not col2:
            col1_str = str(col1).strip()
            # Ignorar cabeçalho
            if 'VOUCHER' not in col1_str.upper():
                # Verificar se parece ser broker (maiúsculas, hífen, etc)
                if col1_str.isupper() or '-' in col1_str or 'AP' in col1_str or 'API' in col1_str:
                    file_brokers.append(col1_str)
                    all_broker_names.add(col1_str)
    
    if file_brokers:
        print(f"Brokers encontrados: {file_brokers}")
    
    wb.close()

print(f"\n{'='*80}")
print("TODOS OS NOMES DE BROKERS ÚNICOS:")
print(f"{'='*80}")
for broker in sorted(all_broker_names):
    print(f"  '{broker}'")
